"""Excel import/export utilities using openpyxl."""

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from fastapi.responses import StreamingResponse


def export_to_excel(
    headers: list[str],
    fields: list[str],
    data: list[dict],
    sheet_name: str = "Sheet1",
) -> StreamingResponse:
    """Export data to an Excel file and return as StreamingResponse.

    Args:
        headers: Column header names (Chinese labels)
        fields: Dict keys to extract from each row
        data: List of dicts to export
        sheet_name: Excel sheet name
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data rows
    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(fields, 1):
            value = item.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    encoded_name = quote(f"{sheet_name}.xlsx")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )


def import_from_excel(file_content: bytes, fields: list[str]) -> list[dict]:
    """Import data from an Excel file.

    Args:
        file_content: Raw bytes of the Excel file
        fields: Field names to map columns to (in order)

    Returns:
        List of dicts with field names as keys
    """
    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header

    result = []
    for row in rows:
        item = {}
        for idx, field in enumerate(fields):
            item[field] = row[idx] if idx < len(row) else None
        result.append(item)

    return result
